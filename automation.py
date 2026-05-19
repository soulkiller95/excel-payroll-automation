import pandas as pd
import os

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer
)

from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.pagesizes import letter

# Create output folder
os.makedirs("output", exist_ok=True)

# Read Excel
df = pd.read_excel("employees.xlsx")

# Calculations
df["Bonus"] = df["Salary"] * 0.10
df["Final Salary"] = df["Salary"] + df["Bonus"]
df["Tax"] = df["Final Salary"] * 0.05
df["Net Salary"] = df["Final Salary"] - df["Tax"]

# Save Excel
excel_output = "output/final_report.xlsx"

df.to_excel(excel_output, index=False)

# Load workbook
wb = load_workbook(excel_output)
ws = wb.active

# Header styles
header_fill = PatternFill(
    start_color="4F81BD",
    end_color="4F81BD",
    fill_type="solid"
)

header_font = Font(
    color="FFFFFF",
    bold=True
)

# Apply header styles
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font

from openpyxl.styles import PatternFill

green_fill = PatternFill(
    start_color="C6EFCE",
    end_color="C6EFCE",
    fill_type="solid"
)

red_fill = PatternFill(
    start_color="FFC7CE",
    end_color="FFC7CE",
    fill_type="solid"
)

# Highlight Net Salary column
for row in range(2, ws.max_row + 1):

    net_salary_cell = ws[f"F{row}"]

    if net_salary_cell.value >= 30000:
        net_salary_cell.fill = green_fill
    else:
        net_salary_cell.fill = red_fill
        
# Auto column widths
for column in ws.columns:

    max_length = 0
    column_letter = column[0].column_letter

    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass

    ws.column_dimensions[column_letter].width = max_length + 2

# Save styled Excel
wb.save(excel_output)

# ---------------- PDF REPORT ---------------- #

pdf_output = "output/payroll_report.pdf"

doc = SimpleDocTemplate(
    pdf_output,
    pagesize=letter
)

elements = []

styles = getSampleStyleSheet()

title = Paragraph(
    "Employee Payroll Report",
    styles["Title"]
)

elements.append(title)
elements.append(Spacer(1, 20))

# Convert dataframe to list
data = [df.columns.tolist()] + df.values.tolist()

# Create table
table = Table(data)

# Table styles
table.setStyle(TableStyle([

    ('BACKGROUND', (0, 0), (-1, 0), colors.blue),
    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),

    ('GRID', (0, 0), (-1, -1), 1, colors.black),

    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),

    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),

    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),

]))

elements.append(table)

# Build PDF
doc.build(elements)

print("Excel and PDF Reports Generated Successfully")